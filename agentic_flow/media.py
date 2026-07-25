from __future__ import annotations

import base64
import csv
import io
import json
import mimetypes
import os
import tempfile
from pathlib import Path
from typing import Any


MAX_ASSET_BYTES = int(os.getenv("AGENTIC_FLOW_MAX_ASSET_MB", "25")) * 1024 * 1024
TEXT_EXTENSIONS = {
    ".txt",
    ".md",
    ".markdown",
    ".csv",
    ".json",
    ".xml",
    ".html",
    ".htm",
    ".yaml",
    ".yml",
    ".log",
    ".rst",
}


def _data_uri(payload: bytes, mime_type: str) -> str:
    return f"data:{mime_type};base64,{base64.b64encode(payload).decode('ascii')}"


def decode_asset(
    value: Any, *, default_name: str = "asset.bin"
) -> tuple[bytes, str, str]:
    name = default_name
    mime_type = ""
    raw = value
    if isinstance(value, dict):
        name = str(value.get("name") or value.get("filename") or default_name)
        mime_type = str(value.get("mime_type") or value.get("content_type") or "")
        raw = value.get("data", value.get("data_uri", value.get("content", "")))
    if isinstance(raw, bytes):
        payload = raw
    elif isinstance(raw, str) and raw.startswith("data:"):
        header, separator, encoded = raw.partition(",")
        if not separator:
            raise ValueError("Data URI inválida.")
        mime_type = header[5:].split(";", 1)[0] or mime_type
        if ";base64" in header:
            payload = base64.b64decode(encoded, validate=True)
        else:
            from urllib.parse import unquote_to_bytes

            payload = unquote_to_bytes(encoded)
    elif isinstance(raw, str):
        try:
            payload = base64.b64decode(raw, validate=True)
        except ValueError:
            payload = raw.encode("utf-8")
    else:
        raise ValueError(
            "O asset precisa ser texto, base64, data URI ou objeto com data/name."
        )
    if len(payload) > MAX_ASSET_BYTES:
        raise ValueError(
            f"Asset excede o limite de {MAX_ASSET_BYTES // (1024 * 1024)} MB."
        )
    if not mime_type:
        mime_type = mimetypes.guess_type(name)[0] or "application/octet-stream"
    return payload, Path(name).name, mime_type


def read_document(value: Any, config: dict[str, Any]) -> dict[str, Any]:
    payload, name, mime_type = decode_asset(value, default_name="document.txt")
    configured_format = str(config.get("format") or "auto").lower()
    extension = (
        f".{configured_format.lstrip('.')}"
        if configured_format != "auto"
        else Path(name).suffix.lower()
    )
    if not extension and mime_type == "application/pdf":
        extension = ".pdf"
    if extension in TEXT_EXTENSIONS or not extension:
        encoding = str(config.get("encoding") or "utf-8")
        text = payload.decode(encoding, errors="replace")
        if extension == ".json":
            text = json.dumps(json.loads(text), ensure_ascii=False, indent=2)
        elif extension == ".csv" and config.get("normalize_csv", False):
            rows = list(csv.reader(io.StringIO(text)))
            text = "\n".join(" | ".join(row) for row in rows)
        pages = None
    elif extension == ".pdf":
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise RuntimeError("Instale pypdf para processar documentos PDF.") from exc
        reader = PdfReader(io.BytesIO(payload))
        text = "\n\n".join(page.extract_text() or "" for page in reader.pages)
        pages = len(reader.pages)
    elif extension == ".docx":
        try:
            from docx import Document
        except ImportError as exc:
            raise RuntimeError("Instale python-docx para processar arquivos DOCX.") from exc
        document = Document(io.BytesIO(payload))
        text = "\n".join(paragraph.text for paragraph in document.paragraphs)
        pages = None
    elif extension == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Instale openpyxl para processar arquivos XLSX.") from exc
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        chunks: list[str] = []
        for sheet in workbook.worksheets:
            chunks.append(f"# {sheet.title}")
            chunks.extend(
                " | ".join("" if cell is None else str(cell) for cell in row)
                for row in sheet.iter_rows(values_only=True)
            )
        text = "\n".join(chunks)
        pages = len(workbook.sheetnames)
    else:
        raise ValueError(
            f"Formato {extension or mime_type} não suportado pelo leitor de documentos."
        )
    max_characters = max(1, int(config.get("max_characters", 200_000)))
    truncated = len(text) > max_characters
    text = text[:max_characters]
    return {
        "text": text,
        "metadata": {
            "name": name,
            "mime_type": mime_type,
            "extension": extension,
            "size_bytes": len(payload),
            "characters": len(text),
            "pages_or_sheets": pages,
            "truncated": truncated,
        },
    }


def process_image(value: Any, config: dict[str, Any]) -> dict[str, Any]:
    try:
        from PIL import Image, ImageOps
    except ImportError as exc:
        raise RuntimeError("Instale Pillow para processar imagens.") from exc
    payload, name, mime_type = decode_asset(value, default_name="image.png")
    with Image.open(io.BytesIO(payload)) as source:
        source.load()
        image = ImageOps.exif_transpose(source)
        original_format = (source.format or "PNG").upper()
        original_size = image.size
        operation = str(config.get("operation") or "inspect")
        if operation == "resize":
            width = max(1, int(config.get("width") or image.width))
            height = max(1, int(config.get("height") or image.height))
            image = ImageOps.contain(image, (width, height))
        elif operation == "grayscale":
            image = ImageOps.grayscale(image)
        elif operation not in {"inspect", "convert"}:
            raise ValueError(f"Operação de imagem desconhecida: {operation}.")
        output_format = str(config.get("output_format") or original_format).upper()
        if output_format == "JPG":
            output_format = "JPEG"
        if output_format not in {"PNG", "JPEG", "WEBP", "GIF", "BMP", "TIFF"}:
            raise ValueError(f"Formato de saída de imagem não suportado: {output_format}.")
        if output_format == "JPEG" and image.mode not in {"RGB", "L"}:
            image = image.convert("RGB")
        output = io.BytesIO()
        save_options = {}
        if output_format in {"JPEG", "WEBP"}:
            save_options["quality"] = min(100, max(1, int(config.get("quality", 90))))
        image.save(output, format=output_format, **save_options)
        encoded = output.getvalue()
        output_mime = Image.MIME.get(output_format, f"image/{output_format.lower()}")
        return {
            "data_uri": _data_uri(encoded, output_mime),
            "metadata": {
                "name": name,
                "source_mime_type": mime_type,
                "format": output_format,
                "mode": image.mode,
                "width": image.width,
                "height": image.height,
                "original_width": original_size[0],
                "original_height": original_size[1],
                "size_bytes": len(encoded),
            },
        }


def extract_video_frames(value: Any, config: dict[str, Any]) -> dict[str, Any]:
    try:
        import cv2
    except ImportError as exc:
        raise RuntimeError(
            "Instale opencv-python-headless para extrair frames de vídeos."
        ) from exc
    payload, name, mime_type = decode_asset(value, default_name="video.mp4")
    suffix = Path(name).suffix or ".mp4"
    temp_path = ""
    capture = None
    try:
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as temporary:
            temporary.write(payload)
            temp_path = temporary.name
        capture = cv2.VideoCapture(temp_path)
        if not capture.isOpened():
            raise ValueError("Não foi possível abrir o vídeo informado.")
        fps = float(capture.get(cv2.CAP_PROP_FPS) or 0)
        total_frames = int(capture.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
        duration = total_frames / fps if fps > 0 else 0
        interval_seconds = max(0.01, float(config.get("interval_seconds", 1)))
        frame_step = max(1, round(fps * interval_seconds)) if fps > 0 else 1
        max_frames = min(100, max(1, int(config.get("max_frames", 12))))
        output_format = str(config.get("output_format") or "jpeg").lower()
        extension = ".png" if output_format == "png" else ".jpg"
        output_mime = "image/png" if output_format == "png" else "image/jpeg"
        frames: list[dict[str, Any]] = []
        index = 0
        while len(frames) < max_frames:
            ok, frame = capture.read()
            if not ok:
                break
            if index % frame_step == 0:
                options = []
                if extension == ".jpg":
                    options = [
                        cv2.IMWRITE_JPEG_QUALITY,
                        min(100, max(1, int(config.get("quality", 88)))),
                    ]
                encoded_ok, buffer = cv2.imencode(extension, frame, options)
                if encoded_ok:
                    frames.append(
                        {
                            "index": index,
                            "timestamp_seconds": round(index / fps, 3)
                            if fps > 0
                            else None,
                            "width": int(frame.shape[1]),
                            "height": int(frame.shape[0]),
                            "data_uri": _data_uri(buffer.tobytes(), output_mime),
                        }
                    )
            index += 1
        return {
            "frames": frames,
            "metadata": {
                "name": name,
                "mime_type": mime_type,
                "fps": fps,
                "total_frames": total_frames,
                "duration_seconds": round(duration, 3),
                "extracted_frames": len(frames),
                "interval_seconds": interval_seconds,
            },
        }
    finally:
        if capture is not None:
            capture.release()
        if temp_path:
            Path(temp_path).unlink(missing_ok=True)
